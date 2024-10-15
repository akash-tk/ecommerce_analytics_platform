from django.test import TestCase
from django.utils import timezone
from analytics_app.models import Order, Product, OrderItem, Customer, Category
from analytics_app.utils import export_monthly_sales_report
from openpyxl import load_workbook
import io

class UtilsTestCase(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='Test Category')
        self.product = Product.objects.create(
            name='Test Product',
            price=100.0,
            category=self.category,
            SKU='test-product-sku'
        )
        self.customer = Customer.objects.create(
            name='Test Customer',
            email='customer@example.com'
        )
        self.order = Order.objects.create(
            customer=self.customer,
            order_date=timezone.now(),
            total_amount=100.0
        )
        self.order_item = OrderItem.objects.create(
            order=self.order,
            product=self.product,
            quantity=5,
            price_at_time_of_order=self.product.price
        )

    def test_export_monthly_sales_report_with_orders(self):
        report = export_monthly_sales_report(month='10', year='2024')
        self.assertIsNotNone(report)
        self.assertEqual(report.status_code, 200)
        self.assertEqual(report['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_export_with_no_orders(self):
        report = export_monthly_sales_report(month='09', year='2024')
        self.assertEqual(report.status_code, 200)
        self.assertEqual(report['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertGreater(len(report.content), 0)

    def test_export_with_invalid_month(self):
        report = export_monthly_sales_report(month='13', year='2024')
        self.assertEqual(report.status_code, 400)

    def test_export_with_empty_report(self):
        report = export_monthly_sales_report(month='02', year='2025')
        self.assertEqual(report.status_code, 200)
        workbook = load_workbook(io.BytesIO(report.content))
        sheet = workbook.active
        self.assertEqual(sheet.max_row, 1)
        self.assertEqual(sheet['A1'].value, "Product")
        self.assertEqual(sheet['B1'].value, "Quantity Sold")
        self.assertEqual(sheet['C1'].value, "Total Revenue")

    def test_export_with_multiple_orders(self):
        Order.objects.create(customer=self.customer, order_date='2024-10-05', total_amount=150.0)
        Order.objects.create(customer=self.customer, order_date='2024-10-15', total_amount=200.0)
        report = export_monthly_sales_report(month='10', year='2024')
        self.assertEqual(report.status_code, 200)
        workbook = load_workbook(io.BytesIO(report.content))
        sheet = workbook.active
        self.assertEqual(sheet['A1'].value, "Product")
        self.assertEqual(sheet['B1'].value, "Quantity Sold")
        self.assertEqual(sheet['C1'].value, "Total Revenue")

    def test_export_with_different_product_quantities(self):
        Order.objects.create(customer=self.customer, order_date='2024-10-10', total_amount=300.0)
        OrderItem.objects.create(order=self.order, product=self.product, quantity=10, price_at_time_of_order=self.product.price)
        report = export_monthly_sales_report(month='10', year='2024')
        self.assertEqual(report.status_code, 200)
        workbook = load_workbook(io.BytesIO(report.content))
        sheet = workbook.active
        self.assertEqual(sheet['B2'].value, 5)
        self.assertEqual(sheet['B3'].value, 10)

    def test_export_monthly_sales_report_invalid_year(self):
        report = export_monthly_sales_report(month='10', year='abc')
        self.assertEqual(report.status_code, 400)
